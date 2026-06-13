import os
import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# 获取当前文件夹下的所有 .xlsx 文件
def get_xlsx_files():
    current_dir = os.getcwd()
    files = [f for f in os.listdir(current_dir) if f.endswith('.xlsx') and not f.endswith('_分类汇总.xlsx')]
    return files

# 根据代码和类别双重匹配进行分流，并标记 Sheet1 中的行
def classify_sheet_data(output_filename):
    wb = load_workbook(output_filename)
    ws_source = wb["Sheet1"]
    
    CLASSIFICATION_MAP = {
        ('72010101020101', '办公用品'): '办公费',
        ('72010101020199', '其它'): '办公费',

        ('72010101020102', '书报杂志'): '印刷费',
        ('720101010202', '印刷费'): '印刷费',

        ('72010101020701', '邮寄费'): '邮电费',
        ('72010101020702', '电话费'): '邮电费',

        ('720101010211', '差旅费'): '国内差旅费',

        ('720101010212', '因公出国（境）费'): '国际差旅费',

        ('7201010102130102', '办公设备维修费'): '维修维护费',
        ('72010101021302', '家具维修费'): '维修维护费',
        ('72010101021303', '公共设施维修费'): '维修维护费', 

        ('72010101021403', '设备租赁'): '租赁费',
        ('72010101021404', '其他租赁'): '租赁费',
        ('72010101023909', '租车费'): '租赁费',
        
        ('72010101019903', '编制外长期聘用人员工资及社保'): '补助劳务费',
        ('72010101022601', '专家咨询费'): '补助劳务费',
        ('72010101022602', '专家评审费'): '补助劳务费',
        ('72010101022699', '其他劳务'): '补助劳务费',
        ('72010101030806', '学生助学金'): '补助劳务费',

        ('72010101023907', '出租车费用'): '交通费',
        
        ('720101010216', '培训费'): '其他',
        ('720101010227', '委托业务费'): '其他',
        ('72010101029999', '其他'): '其他',
        ('900201', '预算数'): '其他',

        ('72010101020703', '网络通讯费'): '软件网络购置费',

        ('72010101100201', '办公家具购置费'): '办公设备购置费',
        ('72010101100202', '一般办公设备购置费'): '办公设备购置费',

        ('72010101021804', '专用工具和仪器'): '其他资本购置费',
        ('72010101040202', '专用设备购置费'): '其他资本购置费'
    }
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_row_idx = 3
    start_row = 5
    max_row = ws_source.max_row
    
    for row_idx in range(start_row, max_row + 1):
        cell_a = ws_source.cell(row=row_idx, column=1).value
        cell_c = ws_source.cell(row=row_idx, column=3).value
        
        if cell_c == "累计发生额" or (cell_a is None and cell_c is None):
            break
            
        code_val = ws_source.cell(row=row_idx, column=4).value
        name_val = ws_source.cell(row=row_idx, column=5).value
        
        code_str = str(code_val).strip() if code_val is not None else ""
        name_str = str(name_val).strip() if name_val is not None else ""
            
        target_sheet_name = None
        if (code_str, name_str) in CLASSIFICATION_MAP:
            target_sheet_name = CLASSIFICATION_MAP[(code_str, name_str)]
        else:
            for (c, n), sheet in CLASSIFICATION_MAP.items():
                if n == name_str:
                    target_sheet_name = sheet
                    break
                    
        if target_sheet_name:
            ws_target = wb[target_sheet_name]
            
            if ws_target.cell(row=1, column=1).value is None:
                for col_idx in range(1, 9):
                    src_header_cell = ws_source.cell(row=header_row_idx, column=col_idx)
                    tgt_header_cell = ws_target.cell(row=1, column=col_idx, value=src_header_cell.value)
                    
                    if src_header_cell.has_style:
                        tgt_header_cell.font = copy(src_header_cell.font)
                        tgt_header_cell.border = copy(src_header_cell.border)
                        tgt_header_cell.fill = copy(src_header_cell.fill)
                        tgt_header_cell.alignment = copy(src_header_cell.alignment)
                        tgt_header_cell.number_format = src_header_cell.number_format
            
            next_row = ws_target.max_row + 1
            if ws_target.max_row == 1 and ws_target.cell(row=1, column=1).value is not None:
                next_row = 2
                
            for col_idx in range(1, 9):
                source_cell = ws_source.cell(row=row_idx, column=col_idx)
                tgt_cell = ws_target.cell(row=next_row, column=col_idx, value=source_cell.value)
                
                if source_cell.has_style:
                    tgt_cell.number_format = source_cell.number_format
                    tgt_cell.alignment = copy(source_cell.alignment)
            
            for col_idx in range(1, 9):
                ws_source.cell(row=row_idx, column=col_idx).fill = yellow_fill

    wb.save(output_filename)

# 对分类表求和，并填入在汇总表中
def add_totals_and_summary(output_filename, prefix, items_sheet_name):
    wb = load_workbook(output_filename)
    sheet_sum_cells = {} 
    
    red_font = Font(color="FF0000")
    center_align = Alignment(horizontal="center", vertical="center")
    
    ws_source = wb["Sheet1"]

    for sheet_name in wb.sheetnames:
        if sheet_name in ["Sheet1", items_sheet_name]:
            continue
            
        ws = wb[sheet_name]
        max_r = ws.max_row
        
        if max_r == 1 and ws.cell(row=1, column=1).value is None:
            for col_idx in range(1, 9):
                src_header_cell = ws_source.cell(row=3, column=col_idx)
                tgt_header_cell = ws.cell(row=1, column=col_idx, value=src_header_cell.value)
                
                if src_header_cell.has_style:
                    tgt_header_cell.font = copy(src_header_cell.font)
                    tgt_header_cell.border = copy(src_header_cell.border)
                    tgt_header_cell.fill = copy(src_header_cell.fill)
                    tgt_header_cell.alignment = copy(src_header_cell.alignment)
                    tgt_header_cell.number_format = src_header_cell.number_format
            
            sheet_sum_cells[sheet_name] = None
            continue
            
        target_row = max_r + 2 if max_r > 1 else 3

        cell_paid_label = ws.cell(row=target_row, column=5, value="已出账")
        cell_paid_label.alignment = center_align
        
        cell_paid_val = ws.cell(row=target_row, column=6)
        if max_r > 1:
            cell_paid_val.value = f"=SUM(F2:F{max_r})"
        else:
            cell_paid_val.value = 0
        cell_paid_val.number_format = '0.00'
        
        cell_unpaid_label = ws.cell(row=target_row + 1, column=5, value="未出账")
        cell_unpaid_label.font = red_font
        cell_unpaid_label.alignment = center_align
        ws.cell(row=target_row + 1, column=6).number_format = '0.00'
        
        cell_total_label = ws.cell(row=target_row + 2, column=5, value="总计")
        cell_total_label.alignment = center_align
        ws.cell(row=target_row + 2, column=6).number_format = '0.00'
        
        sheet_sum_cells[sheet_name] = f"F{target_row}"
    
    ws_summary = wb[items_sheet_name]
    sum_start_row = 1
    
    headers = ["项目", "总计", "已到账", "未到账"]
    for col_idx, h in enumerate(headers, start=1):
        cell_header = ws_summary.cell(row=sum_start_row, column=col_idx, value=h)
        if h == "未到账":
            cell_header.font = red_font
        
    current_row = sum_start_row + 1
    
    for sheet_name, ref_cell in sheet_sum_cells.items():
        ws_summary.cell(row=current_row, column=1, value=sheet_name)
        
        cell_b = ws_summary.cell(row=current_row, column=2, value=f"=C{current_row}+D{current_row}")
        cell_b.number_format = '0.00'
        
        cell_c = ws_summary.cell(row=current_row, column=3)
        if ref_cell is None:
            cell_c.value = 0
        else:
            cell_c.value = f"='{sheet_name}'!{ref_cell}"
        cell_c.number_format = '0.00'
        
        ws_summary.cell(row=current_row, column=4).number_format = '0.00'
        
        current_row += 1
        
    ws_summary.cell(row=current_row, column=1, value="总计")
    
    cell_sum_b = ws_summary.cell(row=current_row, column=2, value=f"=SUM(B{sum_start_row+1}:B{current_row-1})")
    cell_sum_b.number_format = '0.00'
    
    cell_sum_c = ws_summary.cell(row=current_row, column=3, value=f"=SUM(C{sum_start_row+1}:C{current_row-1})")
    cell_sum_c.number_format = '0.00'
    
    cell_sum_d = ws_summary.cell(row=current_row, column=4, value=f"=SUM(D{sum_start_row+1}:D{current_row-1})")
    cell_sum_d.number_format = '0.00'
    
    wb.save(output_filename)

# 主函数：处理财务 Excel 文件
def process_financial_excel():
    files = get_xlsx_files()
    
    if not files:
        print("未找到可以处理的文件")
        return
    
    selected_files = []
    if len(files) == 1:
        selected_files = [files[0]]
    else:
        print("扫描到以下源文件：\n")
        for idx, file in enumerate(files, 1):
            print(f"  [{idx}] {file}")
            
        while True:
            choice = input("\n请输入数字序号选择要处理的文件，多选请用空格隔开，输入 0 处理全部：\n").strip()
            if not choice:
                continue
            
            if choice == '0':
                selected_files = files
                break
                
            try:
                indices = [int(x) - 1 for x in choice.split()]
                if all(0 <= i < len(files) for i in indices):
                    selected_files = [files[i] for i in indices]
                    break
                else:
                    print(f"输入错误，请确保输入的数字在 1 到 {len(files)} 之间")
            except ValueError:
                print("输入错误，请输入正确的数字并用空格隔开")

    for selected_file in selected_files:
        base_name = os.path.splitext(selected_file)[0]
        if base_name.startswith("0000"):
            prefix = base_name[4:]
        else:
            prefix = base_name[:8]
            
        items_sheet_name = f"{prefix}各项"
        output_filename = f"{base_name}_分类汇总.xlsx"

        try:
            shutil.copy(selected_file, output_filename)
        except Exception as e:
            continue

        new_sheet_names = [
            items_sheet_name, "办公费", "印刷费", "邮电费", "国内差旅费",
            "国际差旅费", "维修维护费", "租赁费", "补助劳务费", "交通费",
            "其他", "办公设备购置费", "软件网络购置费", "其他资本购置费"
        ]

        try:
            wb = load_workbook(output_filename)
            for sheet_name in new_sheet_names:
                wb.create_sheet(title=sheet_name)
                    
            wb.save(output_filename)
            
            classify_sheet_data(output_filename)

            add_totals_and_summary(output_filename, prefix, items_sheet_name)

            print(f"{selected_file} 处理完成")
            
        except Exception as e:
            print(f"操作失败: {e}")

if __name__ == "__main__":
    process_financial_excel()